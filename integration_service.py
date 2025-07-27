"""
Integration Service - Модуль интеграций и экспорта данных
Включает в себя интеграции с 1С, CRM, мессенджерами и экспорт в различные форматы
"""

import sqlite3
import json
import pandas as pd
import requests
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import io
import base64
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import zipfile
import os

class IntegrationService:
    """Сервис интеграций с внешними системами"""
    
    def __init__(self, db_path='business_manager.db'):
        self.db_path = db_path
    
    def get_db_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def create_integration(self, user_id: int, company_id: int, integration_type: str, 
                          config: Dict) -> int:
        """Создание новой интеграции"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO integrations (user_id, company_id, integration_type, config_data)
            VALUES (?, ?, ?, ?)
        """, (user_id, company_id, integration_type, json.dumps(config)))
        
        conn.commit()
        integration_id = cursor.lastrowid
        conn.close()
        return integration_id
    
    def get_user_integrations(self, user_id: int, company_id: int = None) -> List[Dict]:
        """Получение интеграций пользователя"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        if company_id:
            cursor.execute("""
                SELECT * FROM integrations 
                WHERE user_id = ? AND company_id = ?
                ORDER BY created_at DESC
            """, (user_id, company_id))
        else:
            cursor.execute("""
                SELECT * FROM integrations 
                WHERE user_id = ?
                ORDER BY created_at DESC
            """, (user_id,))
        
        integrations = cursor.fetchall()
        conn.close()
        
        return [dict(integration) for integration in integrations]
    
    def update_integration_config(self, integration_id: int, config: Dict) -> bool:
        """Обновление конфигурации интеграции"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE integrations 
            SET config_data = ?, last_sync = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (json.dumps(config), integration_id))
        
        conn.commit()
        conn.close()
        return True
    
    def test_integration_connection(self, integration_type: str, config: Dict) -> Dict:
        """Тестирование подключения к интеграции"""
        try:
            if integration_type == '1c':
                return self._test_1c_connection(config)
            elif integration_type == 'crm':
                return self._test_crm_connection(config)
            elif integration_type == 'whatsapp':
                return self._test_whatsapp_connection(config)
            elif integration_type == 'telegram':
                return self._test_telegram_connection(config)
            else:
                return {'success': False, 'message': 'Неизвестный тип интеграции'}
        except Exception as e:
            return {'success': False, 'message': f'Ошибка подключения: {str(e)}'}
    
    def _test_1c_connection(self, config: Dict) -> Dict:
        """Тестирование подключения к 1С"""
        # В демо-режиме возвращаем успешный результат
        return {
            'success': True,
            'message': 'Подключение к 1С успешно',
            'version': '8.3.20.1674',
            'database': config.get('database', 'demo_db')
        }
    
    def _test_crm_connection(self, config: Dict) -> Dict:
        """Тестирование подключения к CRM"""
        # В демо-режиме возвращаем успешный результат
        return {
            'success': True,
            'message': 'Подключение к CRM успешно',
            'crm_type': config.get('crm_type', 'bitrix24'),
            'api_version': 'v1.0'
        }
    
    def _test_whatsapp_connection(self, config: Dict) -> Dict:
        """Тестирование подключения к WhatsApp Business API"""
        # В демо-режиме возвращаем успешный результат
        return {
            'success': True,
            'message': 'Подключение к WhatsApp Business API успешно',
            'phone_number': config.get('phone_number', '+7XXXXXXXXXX')
        }
    
    def _test_telegram_connection(self, config: Dict) -> Dict:
        """Тестирование подключения к Telegram Bot API"""
        # В демо-режиме возвращаем успешный результат
        return {
            'success': True,
            'message': 'Подключение к Telegram Bot успешно',
            'bot_username': config.get('bot_username', '@demo_bot')
        }
    
    def sync_data_from_1c(self, integration_id: int) -> Dict:
        """Синхронизация данных из 1С"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM integrations WHERE id = ?", (integration_id,))
        integration = cursor.fetchone()
        
        if not integration:
            return {'success': False, 'message': 'Интеграция не найдена'}
        
        config = json.loads(integration['config_data'])
        
        # В демо-режиме генерируем тестовые данные
        demo_products = [
            {'name': 'Товар 1С-1', 'price': 1500, 'quantity': 100},
            {'name': 'Товар 1С-2', 'price': 2500, 'quantity': 50},
            {'name': 'Товар 1С-3', 'price': 3500, 'quantity': 25}
        ]
        
        # Добавляем товары в инвентарь
        for product in demo_products:
            cursor.execute("""
                INSERT OR REPLACE INTO inventory (user_id, company_id, name, price, quantity, min_stock)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (integration['user_id'], integration['company_id'], 
                  product['name'], product['price'], product['quantity'], 10))
        
        # Логируем синхронизацию
        cursor.execute("""
            INSERT INTO sync_log (integration_id, sync_type, status, records_processed, completed_at)
            VALUES (?, 'import', 'success', ?, CURRENT_TIMESTAMP)
        """, (integration_id, len(demo_products)))
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'message': f'Синхронизировано {len(demo_products)} товаров из 1С',
            'records_processed': len(demo_products)
        }
    
    def export_data_to_1c(self, integration_id: int) -> Dict:
        """Экспорт данных в 1С"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM integrations WHERE id = ?", (integration_id,))
        integration = cursor.fetchone()
        
        if not integration:
            return {'success': False, 'message': 'Интеграция не найдена'}
        
        # Получаем заказы для экспорта
        cursor.execute("""
            SELECT * FROM orders 
            WHERE user_id = ? AND company_id = ?
            ORDER BY created_at DESC LIMIT 100
        """, (integration['user_id'], integration['company_id']))
        
        orders = cursor.fetchall()
        
        # В демо-режиме просто логируем
        cursor.execute("""
            INSERT INTO sync_log (integration_id, sync_type, status, records_processed, completed_at)
            VALUES (?, 'export', 'success', ?, CURRENT_TIMESTAMP)
        """, (integration_id, len(orders)))
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'message': f'Экспортировано {len(orders)} заказов в 1С',
            'records_processed': len(orders)
        }
    
    def send_whatsapp_message(self, integration_id: int, phone: str, message: str) -> Dict:
        """Отправка сообщения через WhatsApp Business API"""
        # В демо-режиме просто логируем
        print(f"WhatsApp сообщение отправлено на {phone}: {message}")
        
        return {
            'success': True,
            'message': 'Сообщение отправлено через WhatsApp',
            'message_id': f'wamid.{datetime.now().timestamp()}'
        }
    
    def send_telegram_message(self, integration_id: int, chat_id: str, message: str) -> Dict:
        """Отправка сообщения через Telegram Bot API"""
        # В демо-режиме просто логируем
        print(f"Telegram сообщение отправлено в чат {chat_id}: {message}")
        
        return {
            'success': True,
            'message': 'Сообщение отправлено через Telegram',
            'message_id': int(datetime.now().timestamp())
        }

class ExportService:
    """Сервис экспорта данных"""
    
    def __init__(self, db_path='business_manager.db'):
        self.db_path = db_path
    
    def get_db_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def export_to_excel(self, user_id: int, company_id: int, data_type: str, 
                       date_from: str = None, date_to: str = None) -> bytes:
        """Экспорт данных в Excel"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        
        if data_type == 'orders':
            self._export_orders_to_excel(wb, cursor, user_id, company_id, date_from, date_to)
        elif data_type == 'inventory':
            self._export_inventory_to_excel(wb, cursor, user_id, company_id)
        elif data_type == 'customers':
            self._export_customers_to_excel(wb, cursor, user_id, company_id)
        elif data_type == 'financial':
            self._export_financial_to_excel(wb, cursor, user_id, company_id, date_from, date_to)
        
        conn.close()
        
        # Сохраняем в байты
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _export_orders_to_excel(self, wb, cursor, user_id, company_id, date_from, date_to):
        """Экспорт заказов в Excel"""
        ws = wb.active
        ws.title = "Заказы"
        
        # Заголовки
        headers = ['ID', 'Дата', 'Клиент', 'Email', 'Телефон', 'Сумма', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Данные
        query = """
            SELECT id, created_at, customer_name, customer_email, customer_phone, 
                   total_amount, status
            FROM orders 
            WHERE user_id = ? AND company_id = ?
        """
        params = [user_id, company_id]
        
        if date_from:
            query += " AND DATE(created_at) >= ?"
            params.append(date_from)
        if date_to:
            query += " AND DATE(created_at) <= ?"
            params.append(date_to)
        
        query += " ORDER BY created_at DESC"
        
        cursor.execute(query, params)
        orders = cursor.fetchall()
        
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order['id'])
            ws.cell(row=row, column=2, value=order['created_at'][:16])
            ws.cell(row=row, column=3, value=order['customer_name'])
            ws.cell(row=row, column=4, value=order['customer_email'])
            ws.cell(row=row, column=5, value=order['customer_phone'])
            ws.cell(row=row, column=6, value=order['total_amount'])
            ws.cell(row=row, column=7, value=order['status'])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_inventory_to_excel(self, wb, cursor, user_id, company_id):
        """Экспорт склада в Excel"""
        ws = wb.active
        ws.title = "Склад"
        
        # Заголовки
        headers = ['ID', 'Название', 'Цена', 'Количество', 'Мин. остаток', 'Категория']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        cursor.execute("""
            SELECT id, name, price, quantity, min_stock, category
            FROM inventory 
            WHERE user_id = ? AND company_id = ?
            ORDER BY name
        """, (user_id, company_id))
        
        items = cursor.fetchall()
        
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item['id'])
            ws.cell(row=row, column=2, value=item['name'])
            ws.cell(row=row, column=3, value=item['price'])
            ws.cell(row=row, column=4, value=item['quantity'])
            ws.cell(row=row, column=5, value=item['min_stock'])
            ws.cell(row=row, column=6, value=item.get('category', ''))
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_customers_to_excel(self, wb, cursor, user_id, company_id):
        """Экспорт клиентов в Excel"""
        ws = wb.active
        ws.title = "Клиенты"
        
        # Заголовки
        headers = ['ID', 'Имя', 'Email', 'Телефон', 'Адрес', 'Тип', 'LTV', 'Заказов', 'Потрачено']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        cursor.execute("""
            SELECT id, name, email, phone, address, customer_type, ltv, total_orders, total_spent
            FROM customers 
            WHERE company_id = ?
            ORDER BY total_spent DESC
        """, (company_id,))
        
        customers = cursor.fetchall()
        
        for row, customer in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=customer['id'])
            ws.cell(row=row, column=2, value=customer['name'])
            ws.cell(row=row, column=3, value=customer['email'])
            ws.cell(row=row, column=4, value=customer['phone'])
            ws.cell(row=row, column=5, value=customer['address'])
            ws.cell(row=row, column=6, value=customer['customer_type'])
            ws.cell(row=row, column=7, value=customer['ltv'])
            ws.cell(row=row, column=8, value=customer['total_orders'])
            ws.cell(row=row, column=9, value=customer['total_spent'])
    
    def _export_financial_to_excel(self, wb, cursor, user_id, company_id, date_from, date_to):
        """Экспорт финансовых данных в Excel"""
        ws = wb.active
        ws.title = "Финансы"
        
        # Заголовки
        headers = ['ID', 'Дата', 'Тип', 'Категория', 'Сумма', 'Описание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        query = """
            SELECT id, date, type, category, amount, description
            FROM financial_records 
            WHERE user_id = ? AND company_id = ?
        """
        params = [user_id, company_id]
        
        if date_from:
            query += " AND DATE(date) >= ?"
            params.append(date_from)
        if date_to:
            query += " AND DATE(date) <= ?"
            params.append(date_to)
        
        query += " ORDER BY date DESC"
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record['id'])
            ws.cell(row=row, column=2, value=record['date'])
            ws.cell(row=row, column=3, value=record['type'])
            ws.cell(row=row, column=4, value=record['category'])
            ws.cell(row=row, column=5, value=record['amount'])
            ws.cell(row=row, column=6, value=record['description'])
    
    def export_to_csv(self, user_id: int, company_id: int, data_type: str) -> str:
        """Экспорт данных в CSV"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        
        if data_type == 'orders':
            writer.writerow(['ID', 'Дата', 'Клиент', 'Email', 'Сумма', 'Статус'])
            cursor.execute("""
                SELECT id, created_at, customer_name, customer_email, total_amount, status
                FROM orders WHERE user_id = ? AND company_id = ?
            """, (user_id, company_id))
            
            for row in cursor.fetchall():
                writer.writerow([row['id'], row['created_at'], row['customer_name'], 
                               row['customer_email'], row['total_amount'], row['status']])
        
        elif data_type == 'inventory':
            writer.writerow(['ID', 'Название', 'Цена', 'Количество', 'Мин. остаток'])
            cursor.execute("""
                SELECT id, name, price, quantity, min_stock
                FROM inventory WHERE user_id = ? AND company_id = ?
            """, (user_id, company_id))
            
            for row in cursor.fetchall():
                writer.writerow([row['id'], row['name'], row['price'], 
                               row['quantity'], row['min_stock']])
        
        conn.close()
        return csv_buffer.getvalue()
    
    def create_backup_archive(self, user_id: int, company_id: int = None) -> bytes:
        """Создание архива с резервной копией данных"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Создаем ZIP архив в памяти
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Экспортируем каждый тип данных
            data_types = ['orders', 'inventory']
            
            for data_type in data_types:
                try:
                    if company_id:
                        csv_data = self.export_to_csv(user_id, company_id, data_type)
                        zip_file.writestr(f"{data_type}_{company_id}.csv", csv_data)
                    else:
                        # Экспортируем для всех компаний пользователя
                        cursor.execute("SELECT id FROM companies WHERE owner_id = ?", (user_id,))
                        companies = cursor.fetchall()
                        
                        for company in companies:
                            csv_data = self.export_to_csv(user_id, company['id'], data_type)
                            zip_file.writestr(f"{data_type}_company_{company['id']}.csv", csv_data)
                except Exception as e:
                    print(f"Ошибка экспорта {data_type}: {str(e)}")
            
            # Добавляем информацию о резервной копии
            backup_info = {
                'created_at': datetime.now().isoformat(),
                'user_id': user_id,
                'company_id': company_id,
                'version': '1.0'
            }
            zip_file.writestr('backup_info.json', json.dumps(backup_info, indent=2))
        
        conn.close()
        zip_buffer.seek(0)
        return zip_buffer.getvalue()

# Глобальные экземпляры сервисов
integration_service = IntegrationService()
export_service = ExportService()

