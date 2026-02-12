"""
PDS-Ultimate Excel Engine (Standalone)
=========================================
Выделенный движок работы с Excel: создание, чтение, редактирование.

Расширяет встроенный ExcelEngine из file_manager.py:
- Pandas DataFrame ↔ Excel
- Формулы в ячейках
- Множественные листы
- Авто-ширина колонок
- Стилизация: цвета, границы, заголовки
- Chart generation (XlsxWriter charts)

Зависимости: xlsxwriter, openpyxl, pandas (optional)
"""

from __future__ import annotations

import os
from typing import Optional

from pds_ultimate.config import logger


class ExcelEngine:
    """
    Продвинутый Excel-движок.

    Возможности:
    - create(): Создание из structure dict
    - create_from_dataframe(): Из Pandas DataFrame
    - read(): Чтение в dict
    - read_to_dataframe(): Чтение в Pandas DataFrame
    - add_sheet(): Добавление листа в существующий файл
    - add_chart(): Добавление графика
    - set_formulas(): Установка формул
    - auto_width(): Авто-ширина колонок
    """

    # Стандартные форматы
    DEFAULT_HEADER_FORMAT = {
        "bold": True,
        "bg_color": "#4472C4",
        "font_color": "#FFFFFF",
        "border": 1,
        "text_wrap": True,
        "align": "center",
        "valign": "vcenter",
    }

    DEFAULT_CELL_FORMAT = {
        "border": 1,
        "text_wrap": True,
    }

    MONEY_FORMAT = {
        "border": 1,
        "num_format": "$#,##0.00",
    }

    PERCENT_FORMAT = {
        "border": 1,
        "num_format": "0.00%",
    }

    DATE_FORMAT = {
        "border": 1,
        "num_format": "yyyy-mm-dd",
    }

    # ═══════════════════════════════════════════════════════════════════════
    # Create
    # ═══════════════════════════════════════════════════════════════════════

    async def create(
        self,
        filepath: str,
        structure: dict,
        auto_filter: bool = True,
        freeze_header: bool = True,
    ) -> dict:
        """
        Создать Excel из structure dict.

        Structure format:
        {
            "title": "Название",
            "headers": ["Col1", "Col2"],
            "rows": [["val1", "val2"], ...],
            "sheets": [{"name": "Sheet1", "headers": [...], "rows": [...]}],
            "formulas": {"C2": "=A2*B2"},
        }
        """
        try:
            import xlsxwriter

            wb = xlsxwriter.Workbook(filepath)
            header_fmt = wb.add_format(self.DEFAULT_HEADER_FORMAT)
            cell_fmt = wb.add_format(self.DEFAULT_CELL_FORMAT)
            money_fmt = wb.add_format(self.MONEY_FORMAT)

            sheets = structure.get("sheets", [])
            if not sheets:
                sheets = [{
                    "name": structure.get("title", "Данные")[:31],
                    "headers": structure.get("headers", []),
                    "rows": structure.get("rows", []),
                }]

            for sheet_data in sheets:
                ws = wb.add_worksheet(
                    sheet_data.get("name", "Лист")[:31]
                )
                headers = sheet_data.get("headers", [])
                rows = sheet_data.get("rows", [])

                # Заголовки
                for col, h in enumerate(headers):
                    ws.write(0, col, str(h), header_fmt)

                # Авто-ширина
                col_widths = [len(str(h)) + 2 for h in headers]

                # Данные
                for row_idx, row in enumerate(rows, 1):
                    for col_idx, val in enumerate(row):
                        if col_idx >= len(headers):
                            continue

                        # Определить формат
                        try:
                            float_val = float(val)
                            ws.write_number(
                                row_idx, col_idx, float_val, cell_fmt
                            )
                            val_len = len(f"{float_val:.2f}")
                        except (ValueError, TypeError):
                            str_val = str(val or "")
                            ws.write(row_idx, col_idx, str_val, cell_fmt)
                            val_len = len(str_val)

                        if col_idx < len(col_widths):
                            col_widths[col_idx] = max(
                                col_widths[col_idx], val_len + 2
                            )

                # Применить ширину
                for col, width in enumerate(col_widths):
                    ws.set_column(col, col, min(width, 50))

                # Формулы
                formulas = sheet_data.get("formulas", {})
                for cell_ref, formula in formulas.items():
                    ws.write_formula(cell_ref, formula, cell_fmt)

                # Авто-фильтр
                if auto_filter and headers and rows:
                    ws.autofilter(0, 0, len(rows), len(headers) - 1)

                # Закрепление заголовка
                if freeze_header and headers:
                    ws.freeze_panes(1, 0)

            wb.close()

            size = os.path.getsize(filepath) if os.path.exists(filepath) else 0

            return {
                "success": True,
                "path": filepath,
                "format": "xlsx",
                "sheets": len(sheets),
                "size_bytes": size,
            }

        except ImportError:
            return {"error": "xlsxwriter не установлен"}
        except Exception as e:
            return {"error": f"Excel creation failed: {e}"}

    async def create_from_dataframe(
        self,
        filepath: str,
        dataframe,
        sheet_name: str = "Данные",
        index: bool = False,
    ) -> dict:
        """Создать Excel из Pandas DataFrame."""
        try:
            import pandas as pd

            if not isinstance(dataframe, pd.DataFrame):
                return {"error": "Ожидается pandas.DataFrame"}

            structure = {
                "title": sheet_name,
                "headers": list(dataframe.columns),
                "rows": dataframe.values.tolist(),
            }

            if index:
                structure["headers"] = ["Index"] + structure["headers"]
                structure["rows"] = [
                    [str(idx)] + list(row)
                    for idx, row in zip(dataframe.index, dataframe.values)
                ]

            return await self.create(filepath, structure)

        except ImportError:
            return {"error": "pandas не установлен"}
        except Exception as e:
            return {"error": f"DataFrame to Excel failed: {e}"}

    # ═══════════════════════════════════════════════════════════════════════
    # Read
    # ═══════════════════════════════════════════════════════════════════════

    async def read(self, filepath: str) -> dict:
        """
        Прочитать Excel в dict.
        Returns: {"headers": [...], "rows": [...], "sheets": {...}}
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                if ws.max_row and ws.max_row >= 1:
                    headers = [
                        cell.value for cell in ws[1]
                    ]

                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append([v for v in row])

                sheets[sheet_name] = {
                    "headers": headers,
                    "rows": rows,
                    "row_count": len(rows),
                }

            wb.close()

            # Одиночный лист → упрощённый формат
            if len(sheets) == 1:
                data = list(sheets.values())[0]
                data["format"] = "xlsx"
                return data

            return {"sheets": sheets, "format": "xlsx"}

        except ImportError:
            return {"error": "openpyxl не установлен"}
        except Exception as e:
            return {"error": f"Excel read failed: {e}"}

    async def read_to_dataframe(self, filepath: str, sheet_name: Optional[str] = None):
        """Прочитать Excel в Pandas DataFrame."""
        try:
            import pandas as pd
            return pd.read_excel(filepath, sheet_name=sheet_name)
        except ImportError:
            return None
        except Exception as e:
            logger.warning(f"[ExcelEngine] read_to_dataframe failed: {e}")
            return None

    # ═══════════════════════════════════════════════════════════════════════
    # Edit
    # ═══════════════════════════════════════════════════════════════════════

    async def add_column(
        self,
        filepath: str,
        column_name: str,
        values: Optional[list] = None,
        sheet_name: Optional[str] = None,
    ) -> dict:
        """Добавить колонку в существующий Excel."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name] if sheet_name else wb.active

            # Найти следующую свободную колонку
            max_col = ws.max_column + 1

            # Заголовок
            ws.cell(row=1, column=max_col, value=column_name)

            # Значения
            if values:
                for i, val in enumerate(values, 2):
                    ws.cell(row=i, column=max_col, value=val)

            wb.save(filepath)
            wb.close()

            return {
                "success": True,
                "column": column_name,
                "position": max_col,
            }

        except Exception as e:
            return {"error": f"Add column failed: {e}"}

    async def add_row(
        self,
        filepath: str,
        values: list,
        sheet_name: Optional[str] = None,
    ) -> dict:
        """Добавить строку в существующий Excel."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name] if sheet_name else wb.active

            next_row = ws.max_row + 1

            for col, val in enumerate(values, 1):
                ws.cell(row=next_row, column=col, value=val)

            wb.save(filepath)
            wb.close()

            return {
                "success": True,
                "row": next_row,
                "values": len(values),
            }

        except Exception as e:
            return {"error": f"Add row failed: {e}"}

    # ═══════════════════════════════════════════════════════════════════════
    # Charts
    # ═══════════════════════════════════════════════════════════════════════

    async def add_chart(
        self,
        filepath: str,
        chart_type: str = "column",
        data_range: Optional[str] = None,
        title: str = "",
        sheet_name: Optional[str] = None,
    ) -> dict:
        """
        Добавить график в Excel.
        chart_type: column, bar, line, pie, area
        """
        try:
            import openpyxl
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference

            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name] if sheet_name else wb.active

            chart_classes = {
                "column": BarChart,
                "bar": BarChart,
                "line": LineChart,
                "pie": PieChart,
            }

            chart_cls = chart_classes.get(chart_type, BarChart)
            chart = chart_cls()
            chart.title = title or "Данные"

            # Автоматически: первая колонка — категории, остальные — данные
            max_row = ws.max_row
            max_col = ws.max_column

            if max_row > 1 and max_col > 1:
                data = Reference(ws, min_col=2, min_row=1,
                                 max_col=max_col, max_row=max_row)
                cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

            ws.add_chart(chart, f"A{max_row + 3}")

            wb.save(filepath)
            wb.close()

            return {"success": True, "chart_type": chart_type}

        except Exception as e:
            return {"error": f"Chart creation failed: {e}"}

    def get_stats(self) -> dict:
        return {"engine": "excel", "backends": ["xlsxwriter", "openpyxl"]}


# ─── Глобальный экземпляр ────────────────────────────────────────────────────

excel_engine = ExcelEngine()
