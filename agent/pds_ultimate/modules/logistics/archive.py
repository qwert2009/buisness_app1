"""
PDS-Ultimate Archive Manager
================================
Архивация закрытых заказов.

По ТЗ:
1. Все позиции закрытого заказа → копируются в ЕДИНЫЙ АРХИВНЫЙ ФАЙЛ
   (All_Orders_Archive.xlsx) — хранит ВСЕ заказы за ВСЁ время
2. Только после успешного сохранения в архив → временный файл удаляется
3. Итоговые суммы → переносятся в Master_Finance.xlsx
"""

from __future__ import annotations

import os
from datetime import date
from typing import Optional

from pds_ultimate.config import (
    ALL_ORDERS_ARCHIVE_PATH,
    logger,
)


class ArchiveManager:
    """
    Архивариус: архивация в БД + Excel, удаление временных файлов.
    """

    def __init__(self, db_session_factory):
        self._session_factory = db_session_factory

    async def archive_order(self, order_id: int) -> dict:
        """
        Полный цикл архивации заказа:
        1. Копировать позиции в archived_order_items (БД)
        2. Записать в All_Orders_Archive.xlsx
        3. Удалить временный Excel-файл
        4. Установить статус ARCHIVED
        """
        from pds_ultimate.core.database import (
            ArchivedOrderItem,
            Order,
            OrderItem,
            OrderStatus,
        )

        with self._session_factory() as session:
            order = session.query(Order).filter(Order.id == order_id).first()
            if not order:
                return {"error": "Заказ не найден"}

            if order.status not in (OrderStatus.COMPLETED, OrderStatus.ARCHIVED):
                return {"error": f"Заказ в статусе {order.status.value}, нельзя архивировать"}

            items = (
                session.query(OrderItem)
                .filter(OrderItem.order_id == order.id)
                .all()
            )

            # 1. Копирование в архивную таблицу
            supplier_name = order.supplier.name if order.supplier else None
            client_name = order.client.name if order.client else None

            archived_items = []
            for item in items:
                archived = ArchivedOrderItem(
                    original_order_id=order.id,
                    order_number=order.order_number,
                    item_name=item.name,
                    quantity=item.quantity,
                    unit=item.unit,
                    unit_price=item.unit_price,
                    price_currency=item.price_currency,
                    weight=item.weight,
                    tracking_number=item.tracking_number,
                    arrival_date=item.arrival_date,
                    delivery_cost=item.delivery_cost,
                    total_cost=item.total_cost,
                    supplier_name=supplier_name,
                    client_name=client_name,
                    order_income=order.income,
                    order_expense_goods=order.expense_goods,
                    order_delivery_cost=order.delivery_cost,
                    order_net_profit=order.net_profit,
                    order_date=order.order_date,
                    completed_date=order.completed_date,
                    archived_date=date.today(),
                )
                session.add(archived)
                archived_items.append(archived)

            # 2. Запись в Excel-архив
            archive_ok = await self._write_to_archive_excel(
                order, items, supplier_name, client_name
            )

            if not archive_ok:
                session.rollback()
                return {"error": "Ошибка записи в архивный Excel"}

            # 3. Удаление временного файла
            if order.temp_file_path and os.path.exists(order.temp_file_path):
                try:
                    os.remove(order.temp_file_path)
                    logger.info(f"Temp file deleted: {order.temp_file_path}")
                except OSError as e:
                    logger.warning(f"Failed to delete temp file: {e}")

            # 4. Статус ARCHIVED
            order.status = OrderStatus.ARCHIVED
            order.archived_date = date.today()

            session.commit()

            logger.info(
                f"Order #{order.order_number} archived: "
                f"{len(archived_items)} items"
            )

            return {
                "order_number": order.order_number,
                "items_archived": len(archived_items),
                "archive_file": str(ALL_ORDERS_ARCHIVE_PATH),
                "temp_file_deleted": True,
            }

    async def get_archive_stats(self) -> dict:
        """Статистика архива."""
        from pds_ultimate.core.database import ArchivedOrderItem

        with self._session_factory() as session:
            total_items = session.query(ArchivedOrderItem).count()

            # Уникальные заказы
            from sqlalchemy import func
            unique_orders = (
                session.query(func.count(func.distinct(
                    ArchivedOrderItem.order_number
                ))).scalar() or 0
            )

            # Сумма прибыли
            total_profit = (
                session.query(func.sum(ArchivedOrderItem.order_net_profit))
                .filter(ArchivedOrderItem.order_net_profit.isnot(None))
                .scalar() or 0.0
            )

            return {
                "total_orders": unique_orders,
                "total_items": total_items,
                "total_profit": round(total_profit, 2),
                "archive_file": str(ALL_ORDERS_ARCHIVE_PATH),
                "archive_exists": ALL_ORDERS_ARCHIVE_PATH.exists(),
            }

    async def search_archive(
        self,
        query: str,
        limit: int = 50,
    ) -> list[dict]:
        """Поиск в архиве по тексту."""
        from pds_ultimate.core.database import ArchivedOrderItem

        with self._session_factory() as session:
            items = (
                session.query(ArchivedOrderItem)
                .filter(
                    ArchivedOrderItem.order_number.ilike(f"%{query}%")
                    | ArchivedOrderItem.item_name.ilike(f"%{query}%")
                    | ArchivedOrderItem.supplier_name.ilike(f"%{query}%")
                    | ArchivedOrderItem.client_name.ilike(f"%{query}%")
                )
                .order_by(ArchivedOrderItem.archived_date.desc())
                .limit(limit)
                .all()
            )

            return [
                {
                    "order_number": it.order_number,
                    "item_name": it.item_name,
                    "quantity": it.quantity,
                    "supplier": it.supplier_name,
                    "client": it.client_name,
                    "order_income": it.order_income,
                    "net_profit": it.order_net_profit,
                    "archived_date": (
                        it.archived_date.isoformat()
                        if it.archived_date else None
                    ),
                }
                for it in items
            ]

    # ═══════════════════════════════════════════════════════════════════════
    # Excel Archive
    # ═══════════════════════════════════════════════════════════════════════

    async def _write_to_archive_excel(
        self,
        order,
        items: list,
        supplier_name: Optional[str],
        client_name: Optional[str],
    ) -> bool:
        """
        Записать позиции закрытого заказа в All_Orders_Archive.xlsx.
        Если файл существует → добавить строки.
        Если нет → создать с заголовками.
        """
        try:
            import openpyxl

            archive_path = str(ALL_ORDERS_ARCHIVE_PATH)

            if os.path.exists(archive_path):
                wb = openpyxl.load_workbook(archive_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Archive"

                # Заголовки
                headers = [
                    "Заказ", "Дата заказа", "Дата закрытия", "Дата архивации",
                    "Поставщик", "Клиент",
                    "Позиция", "Кол-во", "Ед.", "Цена/ед.", "Валюта",
                    "Трек-номер", "Дата прибытия",
                    "Доставка", "Итого",
                    "Доход заказа", "Расход товар", "Доставка заказа",
                    "Чистая прибыль",
                ]
                ws.append(headers)

                # Стиль заголовков
                from openpyxl.styles import Font, PatternFill
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="4472C4",
                        end_color="4472C4",
                        fill_type="solid",
                    )
                    cell.font = Font(bold=True, color="FFFFFF")

            # Добавление строк
            for item in items:
                row = [
                    order.order_number,
                    order.order_date.isoformat() if order.order_date else "",
                    order.completed_date.isoformat() if order.completed_date else "",
                    date.today().isoformat(),
                    supplier_name or "",
                    client_name or "",
                    item.name,
                    item.quantity,
                    item.unit,
                    item.unit_price,
                    item.price_currency or "USD",
                    item.tracking_number or "",
                    item.arrival_date.isoformat() if item.arrival_date else "",
                    item.delivery_cost,
                    item.total_cost,
                    order.income,
                    order.expense_goods,
                    order.delivery_cost,
                    order.net_profit,
                ]
                ws.append(row)

            wb.save(archive_path)
            logger.info(
                f"Archive Excel updated: {len(items)} rows added "
                f"for order #{order.order_number}"
            )
            return True

        except Exception as e:
            logger.error(f"Failed to write archive Excel: {e}")
            return False

    def format_archive_stats(self, stats: dict) -> str:
        """Форматирование статистики архива."""
        lines = [
            "📁 Архив заказов:",
            f"  📦 Заказов: {stats['total_orders']}",
            f"  📋 Позиций: {stats['total_items']}",
            f"  💰 Общая прибыль: ${stats['total_profit']}",
        ]

        if stats.get("archive_exists"):
            lines.append(f"  📄 Файл: {stats['archive_file']}")
        else:
            lines.append("  ⚠️ Архивный файл ещё не создан")

        return "\n".join(lines)
