"""
PDS-Ultimate Sync Engine
============================
Синхронизация Excel ↔ БД.

По ТЗ (Sync Logic — Файл-как-Эталон):
- Владелец скачивает Excel → правит руками → скидывает обратно
- Бот удаляет старые данные в БД → перезаписывает из файла
- Файл = ЭТАЛОН, БД подстраивается
"""

from __future__ import annotations

import os
from datetime import date, datetime
from typing import Optional

from pds_ultimate.config import MASTER_FINANCE_PATH, logger


class SyncEngine:
    """
    Синхронизация Excel-файлов с БД.
    Направление: Excel → БД (файл = эталон).
    """

    def __init__(self, db_session_factory):
        self._session_factory = db_session_factory

    # ═══════════════════════════════════════════════════════════════════════
    # Синхронизация Master Finance
    # ═══════════════════════════════════════════════════════════════════════

    async def sync_master_finance(self, file_path: Optional[str] = None) -> dict:
        """
        Синхронизировать Master_Finance.xlsx → БД.
        Файл = ЭТАЛОН: удаляем старые данные из БД, загружаем из файла.
        """
        if file_path is None:
            file_path = str(MASTER_FINANCE_PATH)

        if not os.path.exists(file_path):
            return {"error": "Файл не найден"}

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            results = {}

            # Синхронизация листа «Транзакции»
            if "Транзакции" in wb.sheetnames:
                tx_result = await self._sync_transactions_sheet(
                    wb["Транзакции"]
                )
                results["transactions"] = tx_result

            wb.close()

            logger.info(f"Master Finance synced from: {file_path}")
            return {
                "status": "ok",
                "file": file_path,
                "results": results,
            }

        except Exception as e:
            logger.error(f"Sync Master Finance failed: {e}")
            return {"error": str(e)}

    async def sync_archive(self, file_path: Optional[str] = None) -> dict:
        """
        Синхронизировать All_Orders_Archive.xlsx → БД.
        """
        from pds_ultimate.config import ALL_ORDERS_ARCHIVE_PATH

        if file_path is None:
            file_path = str(ALL_ORDERS_ARCHIVE_PATH)

        if not os.path.exists(file_path):
            return {"error": "Архивный файл не найден"}

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            if ws is None:
                return {"error": "Пустой файл"}

            # Читаем заголовки
            headers = [cell.value for cell in ws[1]]
            if not headers:
                return {"error": "Нет заголовков"}

            # Очистить таблицу archived_order_items
            from pds_ultimate.core.database import ArchivedOrderItem

            with self._session_factory() as session:
                session.query(ArchivedOrderItem).delete()

                # Загрузить из файла
                imported = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue

                    data = dict(zip(headers, row))

                    item = ArchivedOrderItem(
                        original_order_id=0,
                        order_number=str(data.get("Заказ", "")),
                        item_name=str(data.get("Позиция", "")),
                        quantity=float(data.get("Кол-во", 0) or 0),
                        unit=str(data.get("Ед.", "шт")),
                        unit_price=self._to_float(data.get("Цена/ед.")),
                        price_currency=str(data.get("Валюта", "USD")),
                        tracking_number=str(data.get("Трек-номер", "") or ""),
                        supplier_name=str(data.get("Поставщик", "") or ""),
                        client_name=str(data.get("Клиент", "") or ""),
                        order_income=self._to_float(data.get("Доход заказа")),
                        order_expense_goods=self._to_float(
                            data.get("Расход товар")
                        ),
                        order_delivery_cost=self._to_float(
                            data.get("Доставка заказа")
                        ),
                        order_net_profit=self._to_float(
                            data.get("Чистая прибыль")
                        ),
                        archived_date=date.today(),
                    )

                    # Парсинг дат
                    order_date_str = data.get("Дата заказа")
                    if order_date_str:
                        item.order_date = self._parse_date(order_date_str)

                    completed_str = data.get("Дата закрытия")
                    if completed_str:
                        item.completed_date = self._parse_date(completed_str)

                    arrival_str = data.get("Дата прибытия")
                    if arrival_str:
                        item.arrival_date = self._parse_date(arrival_str)

                    session.add(item)
                    imported += 1

                session.commit()

            wb.close()

            logger.info(f"Archive synced: {imported} items from {file_path}")
            return {
                "status": "ok",
                "imported": imported,
                "file": file_path,
            }

        except Exception as e:
            logger.error(f"Sync archive failed: {e}")
            return {"error": str(e)}

    # ═══════════════════════════════════════════════════════════════════════
    # Internal: синхронизация транзакций
    # ═══════════════════════════════════════════════════════════════════════

    async def _sync_transactions_sheet(self, ws) -> dict:
        """Синхронизация листа транзакций: файл → БД."""
        from pds_ultimate.core.database import Transaction, TransactionType

        headers = [cell.value for cell in ws[1]]
        if not headers:
            return {"error": "Нет заголовков"}

        type_map = {
            "income": TransactionType.INCOME,
            "expense_goods": TransactionType.EXPENSE_GOODS,
            "expense_delivery": TransactionType.EXPENSE_DELIVERY,
            "expense_personal": TransactionType.EXPENSE_PERSONAL,
            "profit_expenses": TransactionType.PROFIT_EXPENSES,
            "profit_savings": TransactionType.PROFIT_SAVINGS,
        }

        with self._session_factory() as session:
            # Удаляем старые транзакции (файл = эталон)
            session.query(Transaction).delete()

            imported = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                data = dict(zip(headers, row))
                tx_type_str = str(data.get("Тип", "")).lower()
                tx_type = type_map.get(tx_type_str)

                if not tx_type:
                    continue

                tx = Transaction(
                    transaction_type=tx_type,
                    amount=self._to_float(data.get("Сумма")) or 0,
                    currency=str(data.get("Валюта", "USD")),
                    amount_usd=self._to_float(data.get("USD")),
                    description=str(data.get("Описание", "") or ""),
                    category=str(data.get("Категория", "") or ""),
                    transaction_date=(
                        self._parse_date(data.get("Дата")) or date.today()
                    ),
                )

                session.add(tx)
                imported += 1

            session.commit()

        return {"imported": imported}

    # ═══════════════════════════════════════════════════════════════════════
    # Helpers
    # ═══════════════════════════════════════════════════════════════════════

    @staticmethod
    def _to_float(value) -> Optional[float]:
        """Безопасная конвертация в float."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        """Безопасный парсинг даты."""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass

        try:
            return datetime.strptime(str(value), "%d.%m.%Y").date()
        except (ValueError, TypeError):
            pass

        return None
